# -*- coding: utf-8 -*-

# === Force scroll to top on load & reruns ===
import streamlit.components.v1 as _components
def _inject_force_top(_interval_ms: int = 120, _repeat: int = 15, _enable_mo: bool = True) -> None:
    _mo_js = "new MutationObserver(() => { forceTop(); }).observe(document.body, {childList: true, subtree: true});" if _enable_mo else ""
    _html = f"""
    <script>
    (function() {{
      function forceTop() {{
        try {{ window.scrollTo({{top: 0, behavior: 'auto'}}); }} catch(e) {{}}
      }}
      forceTop();
      let ticks = 0;
      const iv = setInterval(() => {{
        forceTop();
        if (++ticks > {{_repeat}}) clearInterval(iv);
      }}, {{_interval_ms}});
      document.addEventListener('visibilitychange', () => {{ if (!document.hidden) forceTop(); }});
      {_mo_js}
    }})();
    </script>
    """
    _components.html(_html, height=0)
# === /Force scroll to top ===
import streamlit as st
import json
import io
import os
import tempfile
import pandas as pd
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

from employment_tax_credit_calc import (
    CompanySize, Region, HeadcountInputs,
    load_params_from_json, calc_gross_credit,
    apply_caps_and_min_tax, calc_clawback, PolicyParameters
)

st.set_page_config(page_title="í†µí•©ê³ ìš©ì¦ëŒ€ ì„¸ì•¡ê³µì œ ê³„ì‚°ê¸°", layout="wide")
# Force scroll to top on load
_inject_force_top()


# =====================
# ìƒë‹¨ ìŠ¤í¬ë¡¤ ê³ ì • (ê°•ì œ)
# =====================
import streamlit.components.v1 as components
components.html(
    """
    <script>
    (function() {
      function forceTop() { try { window.scrollTo({top: 0, behavior: 'auto'}); } catch(e) {} }
      forceTop();
      let ticks = 0;
      const iv = setInterval(() => { forceTop(); if (++ticks > 12) clearInterval(iv); }, 100);
      document.addEventListener('visibilitychange', () => { if (!document.hidden) forceTop(); });
      new MutationObserver(() => { forceTop(); }).observe(document.body, {childList: true, subtree: true});
    })();
    </script>
    """,
    height=0,
)

st.title("í†µí•©ê³ ìš©ì¦ëŒ€ ì„¸ì•¡ê³µì œ ê³„ì‚°ê¸°")
st.caption("ì¡°íŠ¹ë²• Â§29ì¡°ì˜8ì— ë”°ë¥¸ í†µí•©ê³ ìš©ì¦ëŒ€ ì„¸ì•¡ê³µì œë¥¼ ê³„ì‚°í•©ë‹ˆë‹¤.")

# =====================
# ë¡œì»¬ ìºì‹œ ìœ í‹¸
# =====================
def _cache_dir() -> Path:
    try:
        here = Path(__file__).parent
    except NameError:
        here = Path(".").resolve()
    d = here / ".app_cache"
    d.mkdir(parents=True, exist_ok=True)
    return d

def save_cached_logo(png_bytes: bytes):
    try:
        ( _cache_dir() / "logo.png" ).write_bytes(png_bytes)
    except Exception:
        pass

def load_cached_logo() -> bytes | None:
    p = _cache_dir() / "logo.png"
    return p.read_bytes() if p.exists() else None

def save_prefs(company_name: str):
    try:
        pref = {"company_name": company_name}
        ( _cache_dir() / "prefs.json" ).write_text(json.dumps(pref, ensure_ascii=False), encoding="utf-8")
    except Exception:
        pass

def load_prefs() -> dict:
    p = _cache_dir() / "prefs.json"
    if p.exists():
        try:
            return json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}

# =====================
# ì„¸ì…˜ ìƒíƒœ ê¸°ë³¸ ì´ˆê¸°í™” + ìºì‹œ ë¡œë“œ
# =====================
def _ensure(key, default):
    if key not in st.session_state:
        st.session_state[key] = default
    return st.session_state[key]

_ensure("saved_logo_png", None)
_ensure("saved_company_name", None)
_ensure("followup_table", None)
_ensure("calc_summary", None)
_ensure("last_calc", None)

# ìºì‹œì—ì„œ ë¡œê³ /ê¸°ê´€ëª… ë¶ˆëŸ¬ì˜¤ê¸° (ì„¸ì…˜ì´ ë¹„ì–´ ìˆì„ ë•Œë§Œ)
if st.session_state.get("saved_logo_png") is None:
    cached = load_cached_logo()
    if cached:
        st.session_state.saved_logo_png = cached
prefs = load_prefs()
if st.session_state.get("saved_company_name") is None and prefs.get("company_name"):
    st.session_state.saved_company_name = prefs["company_name"]

# ---- rerun ì‹œ NameError ë°©ì§€ìš© ì „ì—­ í”Œë˜ê·¸ ì´ˆê¸°í™” ----
trigger_calc = False

# ==== ì‚¬í›„ê´€ë¦¬ í‘œ ìœ í‹¸ ====
def ensure_followup_table(retention_years:int, default_total:int, default_youth:int):
    import pandas as _pd
    cur = st.session_state.get("followup_table")
    target_years = list(range(1, int(retention_years) + 1))

    if cur is None or getattr(cur, "empty", True):
        st.session_state.followup_table = _pd.DataFrame(
            [{"ì—°ì°¨": y, "ì‚¬í›„ì—°ë„ ìƒì‹œ": int(default_total), "ì‚¬í›„ì—°ë„ ì²­ë…„ë“±": int(default_youth)} for y in target_years]
        )
        return

    cur = cur.copy()
    for col in ["ì—°ì°¨", "ì‚¬í›„ì—°ë„ ìƒì‹œ", "ì‚¬í›„ì—°ë„ ì²­ë…„ë“±"]:
        if col in cur.columns:
            cur[col] = _pd.to_numeric(cur[col], errors="coerce").fillna(0).astype(int)

    map_exist = {int(r["ì—°ì°¨"]): (int(r["ì‚¬í›„ì—°ë„ ìƒì‹œ"]), int(r.get("ì‚¬í›„ì—°ë„ ì²­ë…„ë“±", 0))) for _, r in cur.iterrows()}
    rows = []
    for y in target_years:
        if y in map_exist:
            tot, yth = map_exist[y]
            rows.append({"ì—°ì°¨": y, "ì‚¬í›„ì—°ë„ ìƒì‹œ": tot, "ì‚¬í›„ì—°ë„ ì²­ë…„ë“±": yth})
        else:
            rows.append({"ì—°ì°¨": y, "ì‚¬í›„ì—°ë„ ìƒì‹œ": int(default_total), "ì‚¬í›„ì—°ë„ ì²­ë…„ë“±": int(default_youth)})
    st.session_state.followup_table = _pd.DataFrame(rows).sort_values("ì—°ì°¨").reset_index(drop=True)

with st.sidebar:
    st.header("1) ìµœê·¼ ë²•ë ¹ ì ìš©")
    uploaded = st.file_uploader("ìµœê·¼ ë²•ë ¹ JSON ì—…ë¡œë“œ", type=["json"], accept_multiple_files=False)
    default_info = st.toggle("ì˜ˆì‹œ íŒŒë¼ë¯¸í„° ì‚¬ìš©", value=True)

    st.header("2) ë³´ê³ ì„œ ì˜µì…˜")
    company_name = st.text_input("íšŒì‚¬/ê¸°ê´€ëª… (ë¨¸ë¦¬ê¸€ìš©)", value=st.session_state.saved_company_name or "(ê¸°ê´€ëª…)")
    logo_file = st.file_uploader("íšŒì‚¬ ë¡œê³  (PNG ê¶Œì¥)", type=["png"], accept_multiple_files=False)
    remember_logo = st.checkbox("ì´ ë¡œê³ /ê¸°ê´€ëª…ì„ ê³„ì† ì‚¬ìš©(ì•± ìºì‹œì— ì €ì¥)", value=True)

    logo_bytes = None
    if logo_file is not None:
        logo_bytes = logo_file.getvalue()
        if remember_logo:
            st.session_state.saved_logo_png = logo_bytes
            save_cached_logo(logo_bytes)  # ë””ìŠ¤í¬ ìºì‹œ
    elif st.session_state.saved_logo_png is not None:
        logo_bytes = st.session_state.saved_logo_png or load_cached_logo()

    if company_name:
        st.session_state.saved_company_name = company_name
        if remember_logo:
            save_prefs(company_name)

    params: PolicyParameters = None
    if uploaded is not None:
        try:
            cfg = json.load(uploaded)
            tmp_path = "._tmp_params.json"
            with open(tmp_path, "w", encoding="utf-8") as f:
                json.dump(cfg, f, ensure_ascii=False)
            params = load_params_from_json(tmp_path)
            os.remove(tmp_path)
            st.success("ì—…ë¡œë“œí•œ íŒŒë¼ë¯¸í„°ë¥¼ ë¶ˆëŸ¬ì™”ìŠµë‹ˆë‹¤.")
        except Exception as e:
            st.error(f"íŒŒë¼ë¯¸í„° ë¡œë”© ì‹¤íŒ¨: {e}")
    elif default_info:
        demo_cfg = {
            "per_head_basic": {
                "ì¤‘ì†Œê¸°ì—…": {"ìˆ˜ë„ê¶Œ": 1200000, "ì§€ë°©": 1300000},
                "ì¤‘ê²¬ê¸°ì—…": {"ìˆ˜ë„ê¶Œ": 900000, "ì§€ë°©": 1000000},
                "ëŒ€ê¸°ì—…":   {"ìˆ˜ë„ê¶Œ": 600000, "ì§€ë°©": 700000}
            },
            "per_head_youth": {
                "ì¤‘ì†Œê¸°ì—…": {"ìˆ˜ë„ê¶Œ": 1500000, "ì§€ë°©": 1600000},
                "ì¤‘ê²¬ê¸°ì—…": {"ìˆ˜ë„ê¶Œ": 1100000, "ì§€ë°©": 1200000},
                "ëŒ€ê¸°ì—…":   {"ìˆ˜ë„ê¶Œ": 800000,  "ì§€ë°©": 900000}
            },
            "per_head_conversion": 800000,
            "per_head_return_from_parental": 800000,
            "retention_years": {"ì¤‘ì†Œê¸°ì—…": 3, "ì¤‘ê²¬ê¸°ì—…": 3, "ëŒ€ê¸°ì—…": 2},
            "max_credit_total": None,
            "min_tax_limit_rate": 0.07,
            "excluded_industries": ["ìœ í¥ì£¼ì ì—…", "ê¸°íƒ€ì†Œë¹„ì„±ì„œë¹„ìŠ¤ì—…"]
        }
        tmp_path = "._tmp_params_demo.json"
        with open(tmp_path, "w", encoding="utf-8") as f:
            json.dump(demo_cfg, f, ensure_ascii=False)
        params = load_params_from_json(tmp_path)
        os.remove(tmp_path)
        st.info("ë¯¸ì—…ë¡œë“œì‹œ ì˜ˆì‹œ íŒŒë¼ë¯¸í„°ë¥¼ ì‚¬ìš©)")

st.subheader("ê¸°ì—… ì •ë³´ ë° ì‚¬í›„ê´€ë¦¬ ì˜µì…˜")
colA, colB = st.columns(2)
with colA:
    size_label = st.selectbox("ê¸°ì—…ê·œëª¨", [s.value for s in CompanySize], index=0, key="main_company_size")
    region_label = st.selectbox("ì§€ì—­", [r.value for r in Region], index=1, key="main_region")
    size = CompanySize(size_label)
    region = Region(region_label)
with colB:
    clawback_options = {
        "ë¹„ë¡€ ì¶”ì§• (ê°ì†Œìœ¨ë§Œí¼)": "proportional",
        "ì „ì•¡ ì¶”ì§• (ê°ì†Œ ë°œìƒ ì‹œ ì „ì²´)": "all_or_nothing",
        "êµ¬ê°„ ì¶”ì§• (ê°ì†Œìœ¨ êµ¬ê°„ë³„ ë‹¨ê³„)": "tiered"
    }
    selected_label = st.selectbox("ì¶”ì§• ë°©ì‹ ì„ íƒ", list(clawback_options.keys()), index=0, key="main_clawback_method")
    clawback_method = clawback_options[selected_label]

st.header("ê³ ìš© ì¸ì› ì…ë ¥")
col1, col2, col3 = st.columns(3)
with col1:
    prev_total = st.number_input("ì „ë…„ ìƒì‹œê·¼ë¡œì ìˆ˜", min_value=0, value=50, step=1)
    prev_youth = st.number_input("ì „ë…„ ì²­ë…„ë“± ìƒì‹œê·¼ë¡œì ìˆ˜", min_value=0, value=10, step=1)
with col2:
    curr_total = st.number_input("ë‹¹í•´ ìƒì‹œê·¼ë¡œì ìˆ˜", min_value=0, value=60, step=1)
    curr_youth = st.number_input("ë‹¹í•´ ì²­ë…„ë“± ìƒì‹œê·¼ë¡œì ìˆ˜", min_value=0, value=14, step=1)
with col3:
    converted_regular = st.number_input("ì •ê·œì§ ì „í™˜ ì¸ì› (í•´ë‹¹ì—°ë„)", min_value=0, value=2, step=1)
    returned_parental = st.number_input("ìœ¡ì•„íœ´ì§ ë³µê·€ ì¸ì› (í•´ë‹¹ì—°ë„)", min_value=0, value=1, step=1)

st.header("ì„¸ì•¡ í•œë„/ìµœì €í•œì„¸ ì˜µì…˜")
tax_before_credit = st.number_input("ì„¸ì „ì„¸ì•¡(ìµœì €í•œì„¸ ì ìš© ì‹œ í•„ìš”)", min_value=0, value=120_000_000, step=1)

st.session_state.current_inputs = {
    "company_size": size.value,
    "region": region.value,
    "prev_total": int(prev_total),
    "prev_youth": int(prev_youth),
    "curr_total": int(curr_total),
    "curr_youth": int(curr_youth),
    "converted_regular": int(converted_regular),
    "returned_parental": int(returned_parental),
    "tax_before_credit": int(tax_before_credit),
    "clawback_method": clawback_method,
}

st.divider()
run = st.button("ê³„ì‚°í•˜ê¸°", type="primary", disabled=(params is None))

if run:
    if params is None:
        st.error("íŒŒë¼ë¯¸í„°(JSON)ë¥¼ ë¨¼ì € ë¶ˆëŸ¬ì˜¤ì„¸ìš”.")
        st.stop()

    heads = HeadcountInputs(
        prev_total=int(prev_total),
        curr_total=int(curr_total),
        prev_youth=int(prev_youth),
        curr_youth=int(curr_youth),
        converted_regular=int(converted_regular),
        returned_from_parental_leave=int(returned_parental),
    )
    gross = calc_gross_credit(size, region, heads, params)
    applied = apply_caps_and_min_tax(gross, params, tax_before_credit=int(tax_before_credit) if tax_before_credit else None)
    retention_years = params.retention_years[size]

    st.session_state.calc_summary = {
        "gross": int(gross),
        "applied": int(applied),
        "retention_years": int(retention_years),
        "company_size": size.value,
        "region": region.value,
        "base_headcount": int(curr_total),
        "clawback_method": clawback_method,
    }
    ensure_followup_table(retention_years, int(curr_total), int(curr_youth))

summary = st.session_state.calc_summary
if summary is not None:
    try:
        ensure_followup_table(int(summary["retention_years"]), int(summary["base_headcount"]), int(st.session_state.current_inputs.get("curr_youth", 0)))
    except Exception:
        pass

    st.subheader("â‘  ê³µì œì•¡ ê³„ì‚° ê²°ê³¼")
    st.metric("ì´ê³µì œì•¡ (ìµœì €í•œì„¸/í•œë„ ì „)", f"{summary['gross']:,} ì›")
    st.metric("ì ìš© ê³µì œì•¡ (ìµœì €í•œì„¸/í•œë„ í›„)", f"{summary['applied']:,} ì›")
    st.write(f"ìœ ì§€ê¸°ê°„(ì‚¬í›„ê´€ë¦¬ ëŒ€ìƒ): **{summary['retention_years']}ë…„**")

    st.subheader("â‘¡ ì‚¬í›„ê´€ë¦¬(ì¶”ì§•) ì‹œë®¬ë ˆì´ì…˜ - ë‹¤ë…„í‘œ")
    st.caption("í‘œë¥¼ ì…ë ¥í•œ ë’¤ ì•„ë˜ **[ì¶”ì§•ì„¸ì•¡ ê³„ì‚°í•˜ê¸°]** ë²„íŠ¼ì„ ëˆ„ë¥´ë©´ í‘œê°€ ìë™ ë°˜ì˜ë˜ì–´ ê³„ì‚°ë©ë‹ˆë‹¤.")

    with st.container():
        buf_df = st.session_state.followup_table.copy() if st.session_state.followup_table is not None else pd.DataFrame()
        colcfg = {
            "ì—°ì°¨": st.column_config.NumberColumn("ì—°ì°¨", step=1, disabled=True),
            "ì‚¬í›„ì—°ë„ ìƒì‹œ": st.column_config.NumberColumn("ì‚¬í›„ì—°ë„ ìƒì‹œ", step=1, min_value=0),
            "ì‚¬í›„ì—°ë„ ì²­ë…„ë“±": st.column_config.NumberColumn("ì‚¬í›„ì—°ë„ ì²­ë…„ë“±", step=1, min_value=0),
        }
        edited = st.data_editor(
            buf_df,
            num_rows="fixed",
            hide_index=True,
            key="followup_editor",
            column_config=colcfg,
            use_container_width=True,
        )

    if st.button("ğŸ” ì¶”ì§•ì„¸ì•¡ ê³„ì‚°í•˜ê¸°", type="primary"):
        st.session_state.followup_table = edited.copy()
        trigger_calc = True

    if trigger_calc:
        schedule_records = []
        for _, row in st.session_state.followup_table.iterrows():
            yidx = int(row["ì—°ì°¨"])
            fol_total = int(row["ì‚¬í›„ì—°ë„ ìƒì‹œ"])
            fol_youth = int(row.get("ì‚¬í›„ì—°ë„ ì²­ë…„ë“±", 0))

            claw = calc_clawback(
                credit_applied=int(summary["applied"]),
                base_headcount_at_credit=int(summary["base_headcount"]),
                headcount_in_followup_year=fol_total,
                retention_years_for_company=int(summary["retention_years"]),
                year_index_from_credit=yidx,
                method=summary["clawback_method"],
            )
            schedule_records.append({
                "ì—°ì°¨": yidx,
                "ì‚¬í›„ì—°ë„ ìƒì‹œ": fol_total,
                "ì‚¬í›„ì—°ë„ ì²­ë…„ë“±": fol_youth,
                "ì¶”ì§•ì„¸ì•¡": int(claw),
            })
        schedule_df = pd.DataFrame(schedule_records).sort_values("ì—°ì°¨").reset_index(drop=True)
        total_clawback = int(schedule_df["ì¶”ì§•ì„¸ì•¡"].sum()) if not schedule_df.empty else 0

        st.dataframe(schedule_df, use_container_width=True)
        st.metric("ì¶”ì§•ì„¸ì•¡ í•©ê³„", f"{total_clawback:,} ì›")

        st.session_state.last_calc = {
            **summary,
            "schedule_records": schedule_df.to_dict(orient="records"),
            "total_clawback": total_clawback,
        }

if not trigger_calc:
    _prev = st.session_state.get("last_calc")
    if _prev is not None and _prev.get("schedule_records"):
        import pandas as pd
        schedule_df = pd.DataFrame(_prev["schedule_records"])
        st.subheader("ì‚¬í›„ê´€ë¦¬(ì¶”ì§•) ê²°ê³¼")
        st.dataframe(schedule_df, use_container_width=True)
        st.metric("ì¶”ì§•ì„¸ì•¡ í•©ê³„", f"{int(_prev.get('total_clawback',0)):,} ì›")

safe_total_clawback = (st.session_state.last_calc["total_clawback"]
    if (st.session_state.get("last_calc") and "total_clawback" in st.session_state.last_calc)
    else 0)

st.session_state.calc_context = {
    "company_size": summary["company_size"] if summary else None,
    "region": summary["region"] if summary else None,
    "retention_years": summary["retention_years"] if summary else None,
    "clawback_method": summary["clawback_method"] if summary else None,
    "inputs": st.session_state.get("current_inputs", {}),
    "gross_credit": summary["gross"] if summary else None,
    "applied_credit": summary["applied"] if summary else None,
    "total_clawback": safe_total_clawback,
}

# ============================
# ì—‘ì…€ ìƒì„± (ìš”ì•½ + ì‚¬í›„ê´€ë¦¬ ê²°ê³¼í‘œ) + ìƒë‹¨ ë¡œê³  ì›Œí„°ë§ˆí¬ ì‚½ì… (ì„ì‹œíŒŒì¼ ë°©ì‹)
# ============================
def _build_excel():
    """ì—‘ì…€ ë‚´ë³´ë‚´ê¸°: (1) ê²°ê³¼ìš”ì•½ ì‹œíŠ¸(ìƒë‹¨ ë¡œê³  ì›Œí„°ë§ˆí¬ í¬í•¨), (2) ì‚¬í›„ê´€ë¦¬ ê²°ê³¼í‘œ ì‹œíŠ¸."""
    buffer = io.BytesIO()
    wb = Workbook()
    tmp_logo_path = None

    # ---- ì‹œíŠ¸1: ê²°ê³¼ìš”ì•½ ----
    ws_sum = wb.active
    ws_sum.title = "ê²°ê³¼ìš”ì•½"

    # ë¡œê³  ì›Œí„°ë§ˆí¬ ì‚½ì…: ì„ì‹œíŒŒì¼ì— ì €ì¥ í›„ openpyxlë¡œ ë¡œë“œ (A1 ìœ„ì¹˜)
    start_row = 1
    logo_bytes = st.session_state.get("saved_logo_png") or load_cached_logo()
    if logo_bytes:
        try:
            img = PILImage.open(io.BytesIO(logo_bytes))
            if img.mode != "RGBA":
                img = img.convert("RGBA")
            max_w = 420
            if img.width > max_w:
                ratio = max_w / float(img.width)
                img = img.resize((int(img.width * ratio), int(img.height * ratio)))
            # ë§¤ìš° ì—°í•˜ê²Œ(ì•½ 15% ë¶ˆíˆ¬ëª…)
            r, g, b, a = img.split()
            a = a.point(lambda p: int(p * 0.15))
            img = PILImage.merge("RGBA", (r, g, b, a))
            with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
                img.save(tmp, format="PNG")
                tmp_logo_path = tmp.name
            xl_img = XLImage(tmp_logo_path)
            ws_sum.add_image(xl_img, "A1")
            start_row = 8
            ws_sum.row_dimensions[1].height = 24
        except Exception:
            start_row = 1

    # ë°ì´í„° ì‘ì„±
    summary = st.session_state.get("calc_summary") or {}
    inputs = st.session_state.get("current_inputs") or {}
    last = st.session_state.get("last_calc") or {}

    header_row = start_row
    ws_sum.cell(row=header_row, column=1, value="í•­ëª©")
    ws_sum.cell(row=header_row, column=2, value="ê°’")

    rows = [
        ("ìƒì„±ì¼ì‹œ", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("íšŒì‚¬/ê¸°ê´€ëª…", st.session_state.get("saved_company_name") or ""),
        ("ê¸°ì—…ê·œëª¨", summary.get("company_size", "")),
        ("ì§€ì—­", summary.get("region", "")),
        ("ìœ ì§€ê¸°ê°„(ë…„)", summary.get("retention_years", "")),
        ("ì´ê³µì œì•¡(ìµœì €í•œì„¸/í•œë„ ì „)", f"{summary.get('gross', 0):,}"),
        ("ì ìš© ê³µì œì•¡(ìµœì €í•œì„¸/í•œë„ í›„)", f"{summary.get('applied', 0):,}"),
        ("ì„¸ì „ì„¸ì•¡(ì…ë ¥)", f"{inputs.get('tax_before_credit', 0):,}"),
        ("ì¶”ì§• ë°©ì‹", summary.get("clawback_method", inputs.get('clawback_method', ''))),
        ("ì¶”ì§• í•©ê³„", f"{last.get('total_clawback', 0):,}"),
        ("ì „ë…„ ìƒì‹œ/ì²­ë…„ë“±", f"{inputs.get('prev_total', 0)}/{inputs.get('prev_youth', 0)}"),
        ("ë‹¹í•´ ìƒì‹œ/ì²­ë…„ë“±", f"{inputs.get('curr_total', 0)}/{inputs.get('curr_youth', 0)}"),
        ("ì •ê·œì§ ì „í™˜ / ìœ¡ì•„íœ´ì§ ë³µê·€", f"{inputs.get('converted_regular', 0)} / {inputs.get('returned_parental', 0)}"),
    ]
    r = header_row + 1
    for k, v in rows:
        ws_sum.cell(row=r, column=1, value=k)
        ws_sum.cell(row=r, column=2, value=v)
        r += 1

    bold = Font(bold=True)
    ws_sum.cell(row=header_row, column=1).font = bold
    ws_sum.cell(row=header_row, column=2).font = bold
    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 36

    # ---- ì‹œíŠ¸2: ì‚¬í›„ê´€ë¦¬ ê²°ê³¼í‘œ ----
    ws = wb.create_sheet(title="ì‚¬í›„ê´€ë¦¬ ê²°ê³¼í‘œ")
    headers = ["ì—°ì°¨", "ì‚¬í›„ì—°ë„ ìƒì‹œ", "ì‚¬í›„ì—°ë„ ì²­ë…„ë“±", "ì¶”ì§•ì„¸ì•¡"]
    ws.append(headers)
    last_calc = st.session_state.get("last_calc")
    if last_calc and last_calc.get("schedule_records"):
        for row in last_calc["schedule_records"]:
            ws.append([row["ì—°ì°¨"], row["ì‚¬í›„ì—°ë„ ìƒì‹œ"], row.get("ì‚¬í›„ì—°ë„ ì²­ë…„ë“±", 0), row["ì¶”ì§•ì„¸ì•¡"]])

    try:
        wb.save(buffer)
    finally:
        if tmp_logo_path and os.path.exists(tmp_logo_path):
            try:
                os.remove(tmp_logo_path)
            except Exception:
                pass
    return buffer.getvalue()

excel_bytes = _build_excel()
excel_name = f"tax_credit_result_pro_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
st.download_button(
    label="ì—‘ì…€ ë‹¤ìš´ë¡œë“œ (.xlsx)",
    file_name=excel_name,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    data=excel_bytes,
)

# ==============================
# ğŸ’¬ OpenAI ì±—ë´‡ (ë©”ì¸ í™”ë©´ í•˜ë‹¨) â€” ê¸°ì¡´ êµ¬ì¡° ìœ ì§€
# ==============================
from dotenv import load_dotenv
import importlib, chat_utils
importlib.reload(chat_utils)
from chat_utils import stream_chat

def _build_chat_context() -> str:
    ci = st.session_state.get("current_inputs")
    cc = st.session_state.get("calc_context")
    lines = []
    if ci:
        lines.append(f"[í˜„ì¬ ì…ë ¥] ê¸°ì—…ê·œëª¨={ci.get('company_size')} / ì§€ì—­={ci.get('region')}")
        lines.append(f"ì „ë…„ ìƒì‹œ={ci.get('prev_total')}, ì²­ë…„ë“±={ci.get('prev_youth')} / ë‹¹í•´ ìƒì‹œ={ci.get('curr_total')}, ì²­ë…„ë“±={ci.get('curr_youth')}")
        lines.append(f"ì •ê·œì§ ì „í™˜={ci.get('converted_regular')}, ìœ¡ì•„íœ´ì§ ë³µê·€={ci.get('returned_parental')}")
        if ci.get("tax_before_credit") is not None:
            lines.append(f"ì„¸ì „ì„¸ì•¡={ci.get('tax_before_credit'):,}ì›")
        lines.append(f"ì¶”ì§•ë°©ì‹={ci.get('clawback_method')}")
    if cc:
        lines.append(f"[ìµœê·¼ ê³„ì‚° ê²°ê³¼] ì´ê³µì œì•¡={cc.get('gross_credit'):,}ì› / ì ìš©ê³µì œì•¡={cc.get('applied_credit'):,}ì› / ìœ ì§€ê¸°ê°„={cc.get('retention_years')}ë…„ / ì¶”ì§•í•©ê³„={cc.get('total_clawback'):,}ì›")
    return "\n".join(lines) if lines else ""

load_dotenv()

st.divider()
st.header("ğŸ’¬ OpenAI ì±—ë´‡")
st.caption("ê³„ì‚°ê¸° ì‚¬ìš©ê³¼ ê´€ë ¨í•´ ê¶ê¸ˆí•œ ì ì„ ë¬¼ì–´ë³´ì„¸ìš”.")

if "openai_api_key" not in st.session_state:
    st.session_state.openai_api_key = os.getenv("OPENAI_API_KEY", "")

with st.expander("ğŸ”‘ OpenAI API í‚¤ ì„¤ì •", expanded=not bool(st.session_state.openai_api_key)):
    st.info("ì•„ë˜ì— OpenAI API í‚¤ë¥¼ ì…ë ¥í•˜ì„¸ìš”. (í•œ ë²ˆ ì…ë ¥í•˜ë©´ ì„¸ì…˜ì´ ìœ ì§€ë©ë‹ˆë‹¤.)")
    key_input = st.text_input("API í‚¤ ì…ë ¥ (sk-ë¡œ ì‹œì‘)", type="password", value=st.session_state.openai_api_key)
    if st.button("âœ… ì ìš©í•˜ê¸°", use_container_width=True):
        st.session_state.openai_api_key = key_input.strip()
        if not st.session_state.openai_api_key.startswith("sk-"):
            st.warning("ìœ íš¨í•œ OpenAI API í‚¤ í˜•ì‹ì´ ì•„ë‹™ë‹ˆë‹¤.")
        else:
            os.environ["OPENAI_API_KEY"] = st.session_state.openai_api_key
            st.success("API í‚¤ê°€ ì„¤ì •ë˜ì—ˆìŠµë‹ˆë‹¤. ì´ì œ ì±—ë´‡ì„ ì‚¬ìš©í•  ìˆ˜ ìˆìŠµë‹ˆë‹¤.")

if not st.session_state.openai_api_key:
    st.warning("â›” OpenAI API í‚¤ê°€ ì„¤ì •ë˜ì–´ ìˆì§€ ì•ŠìŠµë‹ˆë‹¤. ìœ„ ì…ë ¥ì°½ì— í‚¤ë¥¼ ì…ë ¥í•˜ì„¸ìš”.")
    st.stop()

if "chat_history" not in st.session_state:
    st.session_state.chat_history = []
if "system_prompt" not in st.session_state:
    st.session_state.system_prompt = "You are a helpful assistant for Korean tax credit calculator users. Reply in Korean by default."

with st.expander("âš™ï¸ ì±—ë´‡ ì„¤ì •", expanded=False):
    model = st.selectbox("ëª¨ë¸ ì„ íƒ", ["gpt-4o-mini", "gpt-4o"], index=0)
    temperature = st.slider("ì˜¨ë„(ì°½ì˜ì„±)", 0.0, 1.0, 0.2, 0.1)
    sys_prompt = st.text_area("ì‹œìŠ¤í…œ í”„ë¡¬í”„íŠ¸", st.session_state.system_prompt, height=80)
    include_ctx = st.checkbox("ì§ˆë¬¸ì— ê³„ì‚° ë§¥ë½ í¬í•¨í•˜ê¸°", value=True)
    apply_pref = st.checkbox("ì„¤ì • ë°˜ì˜í•˜ê¸°", value=True)
    if apply_pref:
        st.session_state.system_prompt = sys_prompt

for m in st.session_state.chat_history:
    with st.chat_message(m["role"]):
        st.markdown(m["content"])

with st.expander("ğŸ ë””ë²„ê·¸(ì´ë²¤íŠ¸ íƒ€ì… í™•ì¸)", expanded=False):
    if st.button("ì´ë²¤íŠ¸ íƒ€ì… ë¯¸ë¦¬ë³´ê¸°"):
        preview = []
        if st.session_state.get("system_prompt"):
            preview.append({"role":"system","type":"input_text"})
        for m in st.session_state.get("chat_history", []):
            role = m.get("role","user")
            typ = "output_text" if role == "assistant" else "input_text"
            preview.append({"role": role, "type": typ})
        st.write(preview if preview else "ì´ë ¥ ì—†ìŒ")

user_text = st.chat_input("ë©”ì‹œì§€ë¥¼ ì…ë ¥í•˜ì„¸ìš”â€¦")
if user_text:
    st.session_state.chat_history.append({"role": "user", "content": user_text})
    with st.chat_message("user"):
        st.markdown(user_text)

    with st.chat_message("assistant"):
        placeholder = st.empty()
        acc = ""
        try:
            ctx = _build_chat_context() if include_ctx else ""
            sys_msg = st.session_state.system_prompt + ("\n\n" + ctx if ctx else "")
            for token in stream_chat(
                st.session_state.chat_history,
                system_prompt=sys_msg,
                model=model,
            ):
                acc += token
                placeholder.markdown(acc)
        except Exception as e:
            acc = f"âš ï¸ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆì–´ìš”: {e}"
            placeholder.markdown(acc)

    st.session_state.chat_history.append({"role": "assistant", "content": acc})
