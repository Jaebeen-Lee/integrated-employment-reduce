# -*- coding: utf-8 -*-
import streamlit as st
import json
import io
import os
import pandas as pd
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

from employment_tax_credit_calc import (
    CompanySize, Region, HeadcountInputs,
    load_params_from_json, calc_gross_credit,
    apply_caps_and_min_tax, calc_clawback, PolicyParameters
)

st.set_page_config(page_title="통합고용세액공제 계산기 (Pro, 메모리 로고·수정)", layout="wide")

st.title("통합고용세액공제 계산기 · Pro (조특법 §29조의8)")
st.caption("로고 메모리 삽입 + 엑셀 서식 적용. NamedStyle 추가 호환성 보완.")

# 세션 상태
if "saved_logo_png" not in st.session_state:
    st.session_state.saved_logo_png = None
if "saved_company_name" not in st.session_state:
    st.session_state.saved_company_name = None

with st.sidebar:
    st.header("1) 정책 파라미터")
    uploaded = st.file_uploader("시행령 기준 파라미터 JSON 업로드", type=["json"], accept_multiple_files=False)
    default_info = st.toggle("예시 파라미터 사용 (업로드 없을 때)", value=True)

    st.header("2) 보고서 옵션")
    company_name = st.text_input("회사/기관명 (머리글용)", value=st.session_state.saved_company_name or "(기관명)")
    logo_file = st.file_uploader("회사 로고 (PNG 권장)", type=["png"], accept_multiple_files=False)
    remember_logo = st.checkbox("이 로고를 계속 사용(세션에 저장)", value=True)

    logo_bytes = None
    if logo_file is not None:
        logo_bytes = logo_file.getvalue()
        if remember_logo:
            st.session_state.saved_logo_png = logo_bytes
    elif st.session_state.saved_logo_png is not None:
        logo_bytes = st.session_state.saved_logo_png

    if company_name and remember_logo:
        st.session_state.saved_company_name = company_name

    params: PolicyParameters = None
    if uploaded is not None:
        try:
            cfg = json.load(uploaded)
            tmp_path = "._tmp_params.json"
            with open(tmp_path, "w", encoding="utf-8") as f:
                json.dump(cfg, f, ensure_ascii=False)
            params = load_params_from_json(tmp_path)
            os.remove(tmp_path)
            st.success("업로드한 파라미터를 불러왔습니다.")
        except Exception as e:
            st.error(f"파라미터 로딩 실패: {e}")
    elif default_info:
        demo_cfg = {
            "per_head_basic": {
                "중소기업": {"수도권": 1200000, "지방": 1300000},
                "중견기업": {"수도권": 900000, "지방": 1000000},
                "대기업":   {"수도권": 600000, "지방": 700000}
            },
            "per_head_youth": {
                "중소기업": {"수도권": 1500000, "지방": 1600000},
                "중견기업": {"수도권": 1100000, "지방": 1200000},
                "대기업":   {"수도권": 800000,  "지방": 900000}
            },
            "per_head_conversion": 800000,
            "per_head_return_from_parental": 800000,
            "retention_years": {"중소기업": 3, "중견기업": 3, "대기업": 2},
            "max_credit_total": None,
            "min_tax_limit_rate": 0.07,
            "excluded_industries": ["유흥주점업", "기타소비성서비스업"]
        }
        tmp_path = "._tmp_params_demo.json"
        with open(tmp_path, "w", encoding="utf-8") as f:
            json.dump(demo_cfg, f, ensure_ascii=False)
        params = load_params_from_json(tmp_path)
        os.remove(tmp_path)
        st.info("예시 파라미터를 사용 중입니다. (업로드 시 자동 대체)")
    # (moved) 기업 정보 & 사후관리 옵션은 본문으로 이동했습니다.

st.subheader("기업 정보 및 사후관리 옵션")
colA, colB = st.columns(2)
with colA:
    size_label = st.selectbox("기업규모", [s.value for s in CompanySize], index=0, key="main_company_size")
    region_label = st.selectbox("지역", [r.value for r in Region], index=1, key="main_region")
    size = CompanySize(size_label)
    region = Region(region_label)
with colB:
    clawback_options = {
        "비례 추징 (감소율만큼)": "proportional",
        "전액 추징 (감소 발생 시 전체)": "all_or_nothing",
        "구간 추징 (감소율 구간별 단계)": "tiered"
    }
    selected_label = st.selectbox("추징 방식 선택", list(clawback_options.keys()), index=0, key="main_clawback_method")
    clawback_method = clawback_options[selected_label]


st.header("고용 인원 입력")
col1, col2, col3 = st.columns(3)

with col1:
    prev_total = st.number_input("전년 상시근로자 수", min_value=0, value=50, step=1)
    prev_youth = st.number_input("전년 청년등 상시근로자 수", min_value=0, value=10, step=1)
with col2:
    curr_total = st.number_input("당해 상시근로자 수", min_value=0, value=60, step=1)
    curr_youth = st.number_input("당해 청년등 상시근로자 수", min_value=0, value=14, step=1)
with col3:
    converted_regular = st.number_input("정규직 전환 인원 (해당연도)", min_value=0, value=2, step=1)
    returned_parental = st.number_input("육아휴직 복귀 인원 (해당연도)", min_value=0, value=1, step=1)

# 현재 입력값을 세션에 저장(챗봇 컨텍스트용)
st.session_state.current_inputs = {
    "company_size": size.value,
    "region": region.value,
    "prev_total": int(prev_total),
    "prev_youth": int(prev_youth),
    "curr_total": int(curr_total),
    "curr_youth": int(curr_youth),
    "converted_regular": int(converted_regular),
    "returned_parental": int(returned_parental),
    "tax_before_credit": int(tax_before_credit) if "tax_before_credit" in locals() else None,
    "clawback_method": clawback_method,
}

st.header("세액 한도/최저한세 옵션")
tax_before_credit = st.number_input("세전세액(최저한세 적용 시 필요)", min_value=0, value=120_000_000, step=1)

st.divider()
run = st.button("계산하기", type="primary", disabled=(params is None))

if run:
    if params is None:
        st.error("파라미터(JSON)를 먼저 불러오세요.")
    else:
        heads = HeadcountInputs(
            prev_total=int(prev_total),
            curr_total=int(curr_total),
            prev_youth=int(prev_youth),
            curr_youth=int(curr_youth),
            converted_regular=int(converted_regular),
            returned_from_parental_leave=int(returned_parental),
        )
        gross = calc_gross_credit(size, region, heads, params)
        applied = apply_caps_and_min_tax(gross, params, tax_before_credit=int(tax_before_credit) if tax_before_credit else None)
        retention_years = params.retention_years[size]

        st.subheader("① 공제액 계산 결과")
        st.metric("총공제액 (최저한세/한도 전)", f"{gross:,} 원")
        st.metric("적용 공제액 (최저한세/한도 후)", f"{applied:,} 원")
        st.write(f"유지기간(사후관리 대상): **{retention_years}년**")
        # 요약 저장 (rerun에도 유지)
        st.session_state.summary = {
            "gross": int(gross),
            "applied": int(applied),
            "retention_years": int(retention_years),
            "curr_total": int(curr_total),
            "curr_youth": int(curr_youth),
        }


        # 메트릭 요약을 세션에 저장해 rerun에도 유지
        st.session_state.summary = {
            "gross": int(gross),
            "applied": int(applied),
            "retention_years": int(retention_years),
            "curr_total": int(curr_total),
            "curr_youth": int(curr_youth),
        }
        # 시뮬레이션 섹션 노출 플래그
        st.session_state.show_sim = True


        # 다년 추징표
        
        
        _render_simulation_pane(params, size, region, clawback_method)

        # 챗봇 컨텍스트로 저장
        st.session_state.calc_context = {
    
            "company_size": size.value,
            "region": region.value,
            "retention_years": int(retention_years),
            "clawback_method": clawback_method,
            "inputs": st.session_state.get("current_inputs", {}),
            "gross_credit": int(gross),
            "applied_credit": int(applied),
            "total_clawback": int(total_clawback),
        }

        # 엑셀 생성
        buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active; ws.title = "Summary"

        # 스타일
        title_font = Font(name="맑은 고딕", size=14, bold=True)
        header_fill = PatternFill("solid", fgColor="F2F2F2")
        thin = Side(style="thin", color="CCCCCC")
        border_all = Border(top=thin, bottom=thin, left=thin, right=thin)
        center = Alignment(horizontal="center", vertical="center")
        right = Alignment(horizontal="right", vertical="center")

        # NamedStyle 등록 (버전 호환: 이미 있으면 예외로 무시)
        currency_style = NamedStyle(name="KRW")
        currency_style.number_format = '#,##0"원"'
        currency_style.alignment = right
        try:
            wb.add_named_style(currency_style)
        except Exception:
            pass

        # 로고 (메모리)
        row_cursor = 1
        if st.session_state.saved_logo_png is not None:
            try:
                pil_img = PILImage.open(io.BytesIO(st.session_state.saved_logo_png))
                img = XLImage(pil_img)
                img.width = 140; img.height = 40
                ws.add_image(img, "A1"); row_cursor = 4
            except Exception as e:
                st.warning(f"로고 삽입 중 오류: {e}")

        title_cell = ws.cell(row=row_cursor, column=1, value="통합고용세액공제 계산 결과")
        title_cell.font = title_font
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=6)
        ws.cell(row=row_cursor, column=7, value=f"작성일자: {datetime.now().strftime('%Y-%m-%d')}").alignment = right
        ws.cell(row=row_cursor+1, column=1, value=f"기관명: {st.session_state.saved_company_name or '(기관명)'}")
        ws.cell(row=row_cursor+1, column=4, value=f"기업규모/지역: {size.value}/{region.value}")

        start = row_cursor + 3
        data = [
            ["항목", "값"],
            ["총공제액 (최저한세/한도 전)", int(gross)],
            ["적용 공제액 (최저한세/한도 후)", int(applied)],
            ["유지기간(년)", int(retention_years)],
            ["추징방식", clawback_method],
            ["추징세액 합계", total_clawback],
        ]
        for r_idx, row in enumerate(data, start=start):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        ws.cell(row=start+1, column=2).style = "KRW"
        ws.cell(row=start+2, column=2).style = "KRW"
        ws.cell(row=start+4, column=2).style = "KRW"

        for r in ws.iter_rows(min_row=start, max_row=start+len(data)-1, min_col=1, max_col=2):
            for cell in r:
                cell.border = border_all
                if cell.row == start:
                    cell.fill = header_fill; cell.alignment = center
                elif cell.column == 1:
                    cell.alignment = center
                else:
                    if cell.style != "KRW":
                        cell.alignment = right

        # 다년 추징표 시트
        ws2 = wb.create_sheet("Clawback Schedule")
        headers = ["연차", "사후연도 인원", "추징세액"]
        ws2.append(headers)
        for row in schedule:
            ws2.append([row["연차"], row["사후연도 인원"], row["추징세액"]])

        for cell in ws2[1]:
            cell.fill = header_fill; cell.border = border_all; cell.alignment = center; cell.font = Font(bold=True)

        for r in range(2, 2 + len(schedule)):
            ws2.cell(row=r, column=1).alignment = center
            ws2.cell(row=r, column=2).alignment = right
            ws2.cell(row=r, column=3).style = "KRW"
            for c in range(1, 4):
                ws2.cell(row=r, column=c).border = border_all

        ws.column_dimensions["A"].width = 22; ws.column_dimensions["B"].width = 26
        for col, w in zip(["A","B","C"], [10, 18, 18]):
            ws2.column_dimensions[col].width = w

        try:
            ws.header_footer.left_header = f"&L{st.session_state.saved_company_name or '(기관명)'}"
            ws.header_footer.right_header = "&R통합고용세액공제 계산 결과"
            ws2.header_footer.left_header = f"&L{st.session_state.saved_company_name or '(기관명)'}"
            ws2.header_footer.right_header = "&RClawback Schedule"
        except Exception:
            pass

        ws3 = wb.create_sheet("Parameters")
        ws3.cell(row=1, column=1, value="Parameters (JSON)")
        ws3.cell(row=2, column=1, value=json.dumps({
            "per_head_basic": {k.value: {kk.value: v for kk, v in d.items()} for k, d in params.per_head_basic.items()},
            "per_head_youth": {k.value: {kk.value: v for kk, v in d.items()} for k, d in params.per_head_youth.items()},
            "per_head_conversion": params.per_head_conversion,
            "per_head_return_from_parental": params.per_head_return_from_parental,
            "retention_years": {k.value: v for k, v in params.retention_years.items()},
            "max_credit_total": params.max_credit_total,
            "min_tax_limit_rate": params.min_tax_limit_rate,
            "excluded_industries": params.excluded_industries,
        }, ensure_ascii=False, indent=2))

        wb.save(buffer)
        excel_name = f"tax_credit_result_pro_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        st.download_button(
            label="엑셀 다운로드 (.xlsx, Pro 포맷)",
            file_name=excel_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            data=buffer.getvalue()
        )

else:
    # rerun으로 버튼 상태가 꺼져도 최근 계산값이 있으면 동일 UI로 다시 표시
    if st.session_state.get("last_calc"):
        _lc = st.session_state["last_calc"]
        st.subheader("① 공제액 계산 결과")
        st.metric("총공제액 (최저한세/한도 전)", f"{_lc['gross']:,} 원")
        st.metric("적용 공제액 (최저한세/한도 후)", f"{_lc['applied']:,} 원")
        st.write(f"유지기간(사후관리 대상): **{_lc['retention_years']}년**")

        st.subheader("② 사후관리(추징) 시뮬레이션 - 다년표")
        import pandas as _pd
        _schedule_df = _pd.DataFrame(_lc["schedule_records"])
        st.dataframe(_schedule_df, use_container_width=True)
        total_clawback = int(_schedule_df["추징세액"].sum())
        st.metric("추징세액 합계", f"{total_clawback:,} 원")
    else:
        if st.session_state.get("summary"):  # 요약이 있으면 언제든 시뮬레이션 표시
            _render_simulation_pane(params, size, region, clawback_method)
        elif st.session_state.get("show_sim") and st.session_state.get("last_calc"): 
            st.subheader("② 사후관리(추징) 시뮬레이션 - 다년표")
            import pandas as _pd
            _df = _pd.DataFrame(st.session_state.last_calc.get("schedule_records", []))
            if not _df.empty:
                st.dataframe(_df, use_container_width=True)
                tc = int(st.session_state.last_calc.get("total_clawback", _df["추징세액"].sum()))
                st.metric("추징세액 합계", f"{tc:,} 원")
        else:
            st.info("좌측에서 파라미터(JSON)를 불러오고, 인원을 입력한 뒤 **계산하기**를 눌러주세요.")



# ---------------------------------------------
# 시뮬레이션 렌더 함수 (요약이 있으면 항상 표/결과 표시)
def _render_simulation_pane(params, size, region, clawback_method):
    import pandas as pd
    st.subheader("② 사후관리(추징) 시뮬레이션 - 다년표")

    # 요약이 없으면 안내
    if "summary" not in st.session_state:
        st.info("먼저 상단에서 **계산하기** 버튼을 눌러 계산 요약을 생성하세요.")
        return

    gross = int(st.session_state.summary["gross"])
    applied = int(st.session_state.summary["applied"])
    retention_years = int(st.session_state.summary["retention_years"])
    curr_total = int(st.session_state.summary["curr_total"])
    curr_youth = int(st.session_state.summary["curr_youth"])

    years = [1, 2, 3]

    # 편집 표를 세션에 보존
    if "sim_df" not in st.session_state or st.session_state.sim_df is None:
        st.session_state.sim_df = pd.DataFrame(
            [{"연차": yr, "사후연도 상시": curr_total, "사후연도 청년등": curr_youth} for yr in years]
        )

    edited = st.data_editor(st.session_state.sim_df, num_rows="fixed", hide_index=True, key="sim_editor")
    st.session_state.sim_df = edited  # 입력해도 사라지지 않도록 세션 반영

    st.caption("연차별 인원을 입력한 뒤 아래 버튼을 눌러 추징세액을 계산하세요.")
    if st.button("🔁 추징세액 계산하기", type="primary", key="btn_clawback"):
        schedule = []
        for _, row in st.session_state.sim_df.iterrows():
            yidx = int(row["연차"])
            fol_total = int(row["사후연도 상시"])
            fol_youth = int(row["사후연도 청년등"]) if "사후연도 청년등" in row else 0

            claw = calc_clawback(
                credit_applied=applied,
                base_headcount_at_credit=curr_total,
                headcount_in_followup_year=fol_total,
                retention_years_for_company=retention_years,
                year_index_from_credit=yidx,
                method=clawback_method,
            )
            schedule.append({"연차": yidx, "사후연도 상시": fol_total, "사후연도 청년등": fol_youth, "추징세액": int(claw)})

        schedule_df = pd.DataFrame(schedule).sort_values("연차").reset_index(drop=True)
        st.dataframe(schedule_df, use_container_width=True)
        total_clawback = int(schedule_df["추징세액"].sum())
        st.metric("추징세액 합계", f"{total_clawback:,} 원")

        # 결과도 세션에 보존 (rerun 유지)
        st.session_state.last_calc = {
            "gross": gross,
            "applied": applied,
            "retention_years": retention_years,
            "company_size": size.value if hasattr(size, "value") else str(size),
            "region": region.value if hasattr(region, "value") else str(region),
            "clawback_method": clawback_method,
            "base_headcount": curr_total,
            "schedule_records": schedule_df.to_dict(orient="records"),
            "total_clawback": total_clawback,
        }
    else:
        # 이전 계산 결과가 있으면 계속 표시
        if st.session_state.get("last_calc") and st.session_state.last_calc.get("schedule_records"):
            _df = pd.DataFrame(st.session_state.last_calc["schedule_records"])
            st.dataframe(_df, use_container_width=True)
            tc = int(st.session_state.last_calc.get("total_clawback", _df["추징세액"].sum()))
            st.metric("추징세액 합계", f"{tc:,} 원")
# ---------------------------------------------
# ==============================
# 💬 OpenAI 챗봇 (메인 화면 하단)
# ==============================
import os
from dotenv import load_dotenv
import importlib, chat_utils
importlib.reload(chat_utils)

def _render_simulation_pane(params, size, region, clawback_method):
    """요약(gross/applied/years)이 있으면 언제든 시뮬레이션 표/결과를 렌더링."""
    import pandas as _pd
    st.subheader("② 사후관리(추징) 시뮬레이션 - 다년표")

    if "summary" not in st.session_state:
        st.info("먼저 상단에서 **계산하기**를 눌러 요약을 생성하세요.")
        return

    gross = st.session_state.summary["gross"]
    applied = st.session_state.summary["applied"]
    retention_years = st.session_state.summary["retention_years"]
    curr_total = st.session_state.summary["curr_total"]
    curr_youth = st.session_state.summary["curr_youth"]

    years = [1, 2, 3]

    # 세션에 편집용 DF가 없으면 생성, 있으면 그대로 사용
    if "sim_df" not in st.session_state or st.session_state.get("sim_df") is None:
        st.session_state.sim_df = _pd.DataFrame(
            [{"연차": yr, "사후연도 상시": int(curr_total), "사후연도 청년등": int(curr_youth)} for yr in years]
        )

    edited = st.data_editor(st.session_state.sim_df, num_rows="fixed", hide_index=True, key="sim_editor_global")
    st.session_state.sim_df = edited  # 편집 내용 유지

    st.caption("연차별 인원을 입력한 후 아래 버튼을 눌러 추징세액을 계산하세요.")
    if st.button("🔁 추징세액 계산하기", type="primary", key="btn_compute_clawback_global"):
        schedule = []
        for _, row in st.session_state.sim_df.iterrows():
            yidx = int(row["연차"])
            fol_total = int(row["사후연도 상시"])
            fol_youth = int(row.get("사후연도 청년등", 0))

            claw = calc_clawback(
                credit_applied=int(applied),
                base_headcount_at_credit=int(curr_total),
                headcount_in_followup_year=fol_total,
                retention_years_for_company=int(retention_years),
                year_index_from_credit=yidx,
                method=clawback_method,
            )
            schedule.append({"연차": yidx, "사후연도 상시": fol_total, "사후연도 청년등": fol_youth, "추징세액": int(claw)})
        schedule_df = _pd.DataFrame(schedule).sort_values("연차").reset_index(drop=True)
        st.dataframe(schedule_df, use_container_width=True)
        total_clawback = int(schedule_df["추징세액"].sum())
        st.metric("추징세액 합계", f"{total_clawback:,} 원")

        # 결과를 세션에 저장
        st.session_state.last_calc = {
            "gross": int(gross),
            "applied": int(applied),
            "retention_years": int(retention_years),
            "company_size": size.value if hasattr(size,"value") else str(size),
            "region": region.value if hasattr(region,"value") else str(region),
            "clawback_method": clawback_method,
            "base_headcount": int(curr_total),
            "schedule_records": schedule_df.to_dict(orient="records"),
            "total_clawback": int(total_clawback),
        }
    else:
        # 이전 결과가 있으면 계속 표시
        if st.session_state.get("last_calc") and st.session_state.last_calc.get("schedule_records"):
            _df = _pd.DataFrame(st.session_state.last_calc["schedule_records"])
            st.dataframe(_df, use_container_width=True)
            tc = int(st.session_state.last_calc.get("total_clawback", _df["추징세액"].sum()))
            st.metric("추징세액 합계", f"{tc:,} 원")
from chat_utils import stream_chat


def _build_chat_context() -> str:
    """현재 입력값과 마지막 계산 결과를 요약해 챗봇에 제공."""
    ci = st.session_state.get("current_inputs")
    cc = st.session_state.get("calc_context")
    lines = []
    if ci:
        lines.append(f"[현재 입력] 기업규모={ci.get('company_size')} / 지역={ci.get('region')}")
        lines.append(f"전년 상시={ci.get('prev_total')}, 청년등={ci.get('prev_youth')} / 당해 상시={ci.get('curr_total')}, 청년등={ci.get('curr_youth')}")
        lines.append(f"정규직 전환={ci.get('converted_regular')}, 육아휴직 복귀={ci.get('returned_parental')}")
        if ci.get("tax_before_credit") is not None:
            lines.append(f"세전세액={ci.get('tax_before_credit'):,}원")
        lines.append(f"추징방식={ci.get('clawback_method')}")
    if cc:
        lines.append(f"[최근 계산 결과] 총공제액={cc.get('gross_credit'):,}원 / 적용공제액={cc.get('applied_credit'):,}원 / 유지기간={cc.get('retention_years')}년 / 추징합계={cc.get('total_clawback'):,}원")
    return "\n".join(lines) if lines else ""
# .env 로드
load_dotenv()

st.divider()
st.header("💬 OpenAI 챗봇")
st.caption("계산기 사용과 관련해 궁금한 점을 물어보세요. (모델: gpt-4o-mini)")

# 🔐 API 키 입력/저장
if "openai_api_key" not in st.session_state:
    st.session_state.openai_api_key = os.getenv("OPENAI_API_KEY", "")

with st.expander("🔑 OpenAI API 키 설정", expanded=not bool(st.session_state.openai_api_key)):
    st.info("아래에 OpenAI API 키를 입력하세요. (한 번 입력하면 세션이 유지됩니다.)")
    key_input = st.text_input("API 키 입력 (sk-로 시작)", type="password", value=st.session_state.openai_api_key)
    if st.button("✅ 적용하기", use_container_width=True):
        st.session_state.openai_api_key = key_input.strip()
        if not st.session_state.openai_api_key.startswith("sk-"):
            st.warning("유효한 OpenAI API 키 형식이 아닙니다.")
        else:
            os.environ["OPENAI_API_KEY"] = st.session_state.openai_api_key
            st.success("API 키가 설정되었습니다. 이제 챗봇을 사용할 수 있습니다.")

# API 키가 없으면 챗봇 비활성화
if not st.session_state.openai_api_key:
    st.warning("⛔ OpenAI API 키가 설정되어 있지 않습니다. 위 입력창에 키를 입력하세요.")
    st.stop()

# 세션 상태 초기화
if "chat_history" not in st.session_state:
    st.session_state.chat_history = []
if "system_prompt" not in st.session_state:
    st.session_state.system_prompt = "You are a helpful assistant for Korean tax credit calculator users. Reply in Korean by default."

# 챗봇 설정
with st.expander("⚙️ 챗봇 설정", expanded=False):
    model = st.selectbox("모델 선택", ["gpt-4o-mini", "gpt-4o"], index=0)
    temperature = st.slider("온도(창의성)", 0.0, 1.0, 0.2, 0.1)
    sys_prompt = st.text_area("시스템 프롬프트", st.session_state.system_prompt, height=80)
    include_ctx = st.checkbox("질문에 계산 맥락 포함하기", value=True)
    apply_pref = st.checkbox("설정 반영하기", value=True)
    if apply_pref:
        st.session_state.system_prompt = sys_prompt

# 대화 이력 표시
for m in st.session_state.chat_history:
    with st.chat_message(m["role"]):
        st.markdown(m["content"])

with st.expander("🐞 디버그(이벤트 타입 확인)", expanded=False):
    if st.button("이벤트 타입 미리보기"):
        # 미리보기용으로 events를 구성해 보여줍니다.
        preview = []
        if st.session_state.get("system_prompt"):
            preview.append({"role":"system","type":"input_text"})
        for m in st.session_state.get("chat_history", []):
            role = m.get("role","user")
            typ = "output_text" if role == "assistant" else "input_text"
            preview.append({"role": role, "type": typ})
        st.write(preview if preview else "이력 없음")

# 입력창
user_text = st.chat_input("메시지를 입력하세요…")
if user_text:
    st.session_state.chat_history.append({"role": "user", "content": user_text})
    with st.chat_message("user"):
        st.markdown(user_text)

    with st.chat_message("assistant"):
        placeholder = st.empty()
        acc = ""
        try:
            ctx = _build_chat_context() if include_ctx else ""
            sys_msg = st.session_state.system_prompt + ("\n\n" + ctx if ctx else "")
            for token in stream_chat(
                st.session_state.chat_history,
                system_prompt=sys_msg,
                model=model,
            ):
                acc += token
                placeholder.markdown(acc)
        except Exception as e:
            acc = f"⚠️ 오류가 발생했어요: {e}"
            placeholder.markdown(acc)

    st.session_state.chat_history.append({"role": "assistant", "content": acc})